import datetime
import random
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction, DataError
from django.db.models import Sum, Q, F, OuterRef
from django.http import QueryDict
from django.shortcuts import render, redirect, reverse
from django.utils import timezone
from django.db.models.functions import ExtractQuarter

from django.http import HttpResponse

from openpyxl import Workbook
from datetime import datetime, timedelta

# Create your views here.
from common.models import Outbound,  User, Customer,Clothes,Man,OutorderClothes
from outbound.forms import OutboundForm, OutorderClothesForm


def list(request):
    form = OutboundForm()
    order = request.GET.get('orderby')  # 获取排序参数

    total_amount = Outbound.objects.aggregate(total=Sum('amount'))['total']
    print(total_amount)
    qs = Outbound.objects.all().order_by('-create_time')
    if order == 'amount':
        qs = qs.order_by('-amount')  # 按照数量
    elif order == 'man':
        qs = qs.order_by('-amount')  # 按照领取人
    elif order == 'address':
        qs = qs.order_by('customer__address')  # 按照车号
    elif order == 'clothes':
        qs = qs.order_by('clothes__name')  # 按照机件
    elif order == 'jisuan':
        qs = qs.annotate(product_price=F('clothes__price') * F('amount')).order_by('-product_price')
    total_cheng = 0
    total_price = 0  # 初始化总价为 0
    for inbound in qs:
        price = inbound.clothes.price
        amount = inbound.amount
        product_price = price * amount
        total_price += product_price
        product_price = round(product_price, 2)
        inbound.product_price = product_price
        total_cheng += inbound.product_price

    paginator = Paginator(qs, 50)
    page = request.GET.get('page', '1')
    result = paginator.page(page)

    context = {
        'result': result,
        'total_amount': total_amount,
        'total_price':total_cheng,
        'outbound_form': form,

    }
    return render(request,'outbound/index.html', context)
    return search(request, clothes=form.cleaned_data['clothes'], context=context)


def add(request):
    latest = Outbound.objects.last()
    print('数据:', latest)
    if request.method == "POST":
        outbound_form = OutboundForm(request.POST)
        time = request.POST.get('start_time')
        if outbound_form.is_valid():
            costomer = outbound_form.cleaned_data['customer']
            uid = request.session.get('user_id')
            user = User.objects.get(id=uid)
            now = datetime.now().strftime('%Y%m%d%H%M')
            code = 'OUT' + now
            name = outbound_form.cleaned_data['name']
            clothes = outbound_form.cleaned_data['clothes']
            amount = outbound_form.cleaned_data['amount']
            print(time)
            if time:
                with transaction.atomic():  # 事务
                    # 库存足够才成功
                    if clothes.stock >= amount:
                        # 出库减少库存
                        clothes.stock -= amount
                        clothes.save()

                        new_outbound = Outbound.objects.create(code=code,
                                                   customer=costomer,
                                                   user=user,
                                                   name=name,
                                                   clothes=clothes,
                                                   amount=amount,create_time= timezone.datetime.strptime(time, '%Y-%m-%d'))
                        print(new_outbound.create_time)
                        context = {
                            'id': new_outbound.id
                        }
                        messages.add_message(request, messages.SUCCESS, '添加成功')
                        return redirect(reverse('outbound:index', ))

                    else:
                        context = {
                            'outbound_form': outbound_form
                        }
                        messages.add_message(request, messages.WARNING, '请检查填写的内容')
                        return render(request, 'outbound/add.html', context)
            else:
                with transaction.atomic():  # 事务
                    # 库存足够才成功
                    if clothes.stock >= amount:
                        # 出库减少库存
                        clothes.stock -= amount
                        clothes.save()
                        new_outbound = Outbound.objects.create(code=code,
                                                               customer=costomer,
                                                               user=user,
                                                               name=name,
                                                               clothes=clothes,
                                                               amount=amount,
                                                               )
                        context = {
                            'id': new_outbound.id
                        }
                        messages.add_message(request, messages.SUCCESS, '添加成功')
                        return redirect(reverse('outbound:index'))

                    else:
                        context = {
                            'outorder_form': outbound_form,

                        }
                        messages.add_message(request, messages.WARNING, '添加失败，库存不足')
                        return render(request, 'outbound/add.html', context)
        else:


            context = {
                'outbound_form': outbound_form
            }
            messages.add_message(request, messages.WARNING, '请检查填写的内容')
            return render(request, 'outbound/add.html', context)
    else:
        outorder_form = OutboundForm(initial={'customer': latest.customer,
                'name': latest.name,
                'clothes': latest.clothes,})  # 将最近一次提交的数据传递给表单的initial参数

        #outorder_form = OutboundForm()

        context = {
            'outorder_form': outorder_form,

        }
        return render(request, 'outbound/add.html', context)


def search(request, clothes=None, context=None):
    form = OutboundForm(initial={'adminuser':'3'})
    if request.method == "GET":

        outbound_objects = Outbound.objects
        qs = outbound_objects.all()
        #qs = outbound_objects.select_related('clothes').all()
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
        address = request.GET.get('address')
        clothes_name = request.GET.get('clothes_name')
        clothes_id = request.GET.get('clothes_id')
        clothes_sn = request.GET.get('clothes_sn')
        man_name = request.GET.get('user_name')
        print('搜索的:',man_name)
        # 构建查询条件
        conditions = Q()

        if start_time and end_time:
            conditions &= Q(create_time__range=(start_time, end_time))

        if address:
            conditions &= Q(customer__address=address)

        if clothes_name:
            conditions &= Q(clothes__name=clothes_name)

        if clothes_sn:
            conditions &= Q(clothes__sn=clothes_sn)
        if clothes_id:
            conditions &= Q(clothes_id=clothes_id)
        if man_name:
            conditions &= Q(name__name=man_name)
        print('显示:',conditions)

        # 应用查询条件
        qs = qs.filter(conditions).order_by('-amount')

        total_amount = qs.aggregate(total=Sum('amount'))['total']
        #total_price = qs.aggregate(total=Sum('clothes__price'))['total']
        total_price = 0
        for inbound in qs:
            price = inbound.clothes.price
            amount = inbound.amount
            product_price = price * amount
            total_price += product_price
            product_price = round(product_price, 2)
            inbound.product_price = product_price


        paginator = Paginator(qs, 100)
        page = request.GET.get('page', 1)
        result = paginator.page(page)

        context = {
            'result': result,
            'total_amount': total_amount,
            'total_price': total_price,
            'outbound_form': form,


        }

        messages.add_message(request, messages.SUCCESS, '查询成功')
        return render(request, 'outbound/index.html', context)
    else:
        form = OutboundForm(initial={'adminuser':'3'})
        context = {
            'outorder_form': object,
            'outbound_form': form,
        }

        messages.add_message(request, messages.SUCCESS, '查询失败')
        return render(request, 'outbound/index.html', context)


def update(request, outbound_id):
    outbound = Outbound.objects.get(id=outbound_id)
    if request.method == "POST":
        outbound_form = OutboundForm(request.POST)
        if outbound_form.is_valid():
            customer = outbound_form.cleaned_data['customer']
            user = outbound_form.cleaned_data['user']
            create_time = outbound_form.cleaned_data['create_time']
            name = outbound_form.cleaned_data['name']
            amount = outbound_form.cleaned_data['amount']
            del_amount = amount - outbound.amount
            if del_amount <= outbound.clothes.stock:
                with transaction.atomic():
                    outbound.amount = amount
                    outbound.clothes.stock -= del_amount

                    if customer:
                        outbound.customer = customer
                    if create_time:
                        outbound.create_time = create_time
                    if user:
                        outbound.user = user
                    if name:
                        outbound.name = name

                    outbound.save()
                    outbound.clothes.save()
            context = {
                'outorder_id': outbound_id
            }
            messages.add_message(request, messages.SUCCESS, '修改成功')
            return redirect(reverse('outbound:index'))

        else:
            context = {
                'outorder_form': outbound_form,

            }
            messages.add_message(request, messages.WARNING, '请检查填写的内容')
            return render(request, 'outbound/edit.html', context)
    else:
        outorder_form = OutboundForm({'id': outbound.id,
                                      'code': outbound.code,
                                      'customer': outbound.customer,
                                      'amount': outbound.amount,
                                      'create_time': outbound.create_time,
                                      'name': outbound.name,
                                      'clothes': outbound.clothes,
                                      'user': outbound.user,


                                      })
        context = {
            'outorder_form': outorder_form,
            'outbound_id': outbound_id,

        }
        return render(request, 'outbound/edit.html', context)


def delete(request, outbound_id):
    outbound = Outbound.objects.get(id=outbound_id)
    with transaction.atomic():
        outbound.delete()
        outbound.clothes.stock += outbound.amount
        outbound.clothes.save()
    messages.add_message(request, messages.SUCCESS, '删除成功')
    return redirect(reverse('outbound:index'))



def addmore(request, outorder_id):
    if request.method == "POST":
        time = request.POST.get('start_time')
        print('时间',time)
        outorderclothes_form = OutorderClothesForm(request.POST)
        if outorderclothes_form.is_valid():
            clothes = outorderclothes_form.cleaned_data['clothes']
            amount = outorderclothes_form.cleaned_data['amount']
            name = outorderclothes_form.cleaned_data['name']


            if time:
                with transaction.atomic():  # 事务
                    # 库存足够才成功
                    if clothes.stock >= amount:
                        # 出库减少库存
                        clothes.stock -= amount
                        clothes.save()
                        new_outorderclothes = OutorderClothes.objects.create(clothes=clothes,outorder_id=outorder_id,amount=amount,name=name,create_time=timezone.datetime.strptime(time, '%Y-%m-%d'))
                        print('时间查看',new_outorderclothes)
                        context = {
                            'id': new_outorderclothes.id
                        }
                        messages.add_message(request, messages.SUCCESS, '添加成功')
                        return redirect(reverse('outorder:detail', args={outorder_id}))

                    else:
                        context = {
                            'outorderclothes_form': outorderclothes_form,
                            'outorder_id': outorder_id
                        }
                        messages.add_message(request, messages.WARNING, '添加失败，库存不足')
                        return render(request, 'outorder/addmore.html', context)
            else:
                with transaction.atomic():  # 事务
                    # 库存足够才成功
                    if clothes.stock >= amount:
                        # 出库减少库存
                        clothes.stock -= amount
                        clothes.save()
                        new_outorderclothes = OutorderClothes.objects.create(clothes=clothes, outorder_id=outorder_id,
                                                                             amount=amount, name=name)
                        context = {
                            'id': new_outorderclothes.id
                        }
                        messages.add_message(request, messages.SUCCESS, '添加成功')
                        return redirect(reverse('outorder:detail', args={outorder_id}))

                    else:
                        context = {
                            'outorderclothes_form': outorderclothes_form,
                            'outorder_id': outorder_id
                        }
                        messages.add_message(request, messages.WARNING, '添加失败，库存不足')
                        return render(request, 'outorder/addmore.html', context)
        else:
            context = {
                'outorderclothes_form': outorderclothes_form,
                'outorder_id': outorder_id
            }
            messages.add_message(request, messages.WARNING, '请检查填写的内容')
            return render(request, 'outorder/addmore.html', context)
    else:
        outorderclothes_form = OutorderClothesForm()
        shuju = Clothes.objects.all()
        context = {
            'outorderclothes_form': outorderclothes_form,
            'outorder_id': outorder_id,
            'shuju': shuju,
        }
        return render(request, 'outorder/addmore.html', context)


def asmore(request, outorder_id):
    if request.method == "POST":
        shuju = Clothes.objects.all()
        outorderclothes_form = OutorderClothesForm(request.POST)
        if outorderclothes_form.is_valid():
            clothes = outorderclothes_form.cleaned_data['clothes']
            amount = outorderclothes_form.cleaned_data['amount']
            name = outorderclothes_form.cleaned_data['name']
            create_time = outorderclothes_form.cleaned_data['create_time']
            now = datetime.datetime.now().strftime('%Y%m%d%H%M')

            with transaction.atomic():  # 事务
                # 库存足够才成功
                if clothes.stock >= amount:
                    # 出库减少库存
                    clothes.stock -= amount
                    clothes.save()
                    new_outorderclothes = OutorderClothes.objects.create(clothes=clothes, outorder_id=outorder_id, amount=amount,name=name,
 )
                    context = {
                        'id': new_outorderclothes.id
                    }
                    messages.add_message(request, messages.SUCCESS, '添加成功')
                    return redirect(reverse('outorder:detail', args={outorder_id}))

                else:
                    context = {
                        'outorderclothes_form': outorderclothes_form,
                        'outorder_id': outorder_id
                    }
                    messages.add_message(request, messages.WARNING, '添加失败，库存不足')
                    return render(request, 'outorder/addmore.html', context)

        else:
            context = {
                'outorderclothes_form': outorderclothes_form,
                'outorder_id': outorder_id
            }
            messages.add_message(request, messages.WARNING, '请检查填写的内容')
            return render(request, 'outorder/addmore.html', context)
    else:
        outorderclothes_form = OutorderClothesForm()
        shuju = Clothes.objects.all()
        context = {
            'outorderclothes_form': outorderclothes_form,
            'outorder_id': outorder_id,
            'shuju': shuju,
        }
        return render(request, 'outorder/addmore.html', context)
def detail(request, outbound_id):

    # 返回一个 QuerySet 对象 ，包含所有的表记录

    qs2 = Outbound.objects.filter(outbound_id=outbound_id)


    sum = 0
    shuliang = 0
    for foo in qs2:
        sum += foo.clothes.price * foo.amount
        shuliang += foo.amount
        print("总和:", sum)
        print(shuliang)
    context = {

        'qs2': qs2,
        'sum': sum,
        'shuliang':shuliang,
        'outorder_id': outbound_id
    }
    return render(request, 'outbound/detail.html', context)
from django.http.response import JsonResponse, HttpResponseBadRequest


def test(request):
    # 输入车号，获取本机台机件应用情况。
    address = '03#'
    obj = Outbound.objects.filter(customer__address=address)
    print(obj)
    return HttpResponse("OK")
def get_total_amount_per_quarter():
    # 使用 annotate 和 values 方法进行聚合和分组
    queryset = Outbound.objects.annotate(
        quarter=ExtractQuarter('create_time')
    ).values('customer__address', 'quarter').annotate(
        total_amount = Sum(F('amount') * F('clothes__price'))
    ).order_by('-total_amount')

    return queryset
def your_view_function(request):
    quarter = request.GET.get('quarter')
    if quarter not in ['1', '2', '3', '4']:
        error_message = "日期不正确"
        return render(request, 'outbound/biao.html', {'error_message': error_message})

    if not quarter:
        current_date = timezone.now().date()
        current_quarter = (current_date.month - 1) // 3 + 1
        quarter = str(current_quarter)

    totals = get_total_amount_per_quarter().filter(quarter=quarter)
    return render(request, 'outbound/biao.html', {'result': totals})
def download(request):
    totals = get_total_amount_per_quarter()  # 调用你的聚合函数获取数据

    # 生成 Excel 文件
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Totals'

    # 写入表头
    worksheet['A1'] = '地址'
    worksheet['B1'] = '季度'
    worksheet['C1'] = '总金额'

    # 写入数据
    for index, total in enumerate(totals, start=2):
        worksheet.cell(row=index, column=1, value=total['customer__address'])
        worksheet.cell(row=index, column=2, value=total['quarter'])
        worksheet.cell(row=index, column=3, value=total['total_amount'])

    # 保存 Excel 文件到内存中
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=totals.xlsx'
    workbook.save(response)

    return response

from django.db.models import Sum
from django.shortcuts import render
from django.utils import timezone
from datetime import datetime, timedelta

def line_chart_view(request):
    # 获取当前日期
    current_date = timezone.now().date()

    # 创建一个列表来存储每个月的消耗总价
    monthly_totals = []

    # 循环遍历过去12个月的日期
    for i in range(1, 13):
        # 计算每个月的起始日期和结束日期
        month = current_date.month - i
        year = current_date.year

        # 处理月份小于1的情况
        if month <= 0:
            month += 12
            year -= 1

        start_date = datetime(year, month, 1)
        end_date = start_date.replace(day=1, month=month % 12 + 1, year=year) - timedelta(days=1)

        # 计算每个月的消耗总价
        monthly_total = Outbound.objects.filter(create_time__range=(start_date, end_date)).aggregate(total=Sum('clothes__price'))['total']
        print(monthly_total)
        monthly_totals.append(monthly_total or 0)  # 如果没有消耗数据，则将总价设为0
        monthly_totals_str = [str(total) for total in monthly_totals]

    return render(request, 'outbound/line_chart.html', {'monthly_totals': monthly_totals_str})
def zdlist(request):



    address_amount_totals = Outbound.objects.values('customer__address') \
        .annotate(total_amount=Sum('amount')) \
        .order_by('-total_amount')

    for address_total in address_amount_totals:
        address = address_total['customer__address']
        total_amount = address_total['total_amount']
        print(f"车号: {address} - 数量: {total_amount}")

    sn_amount_totals = Outbound.objects.values('clothes__sn') \
        .annotate(total_amount=Sum('amount')) \
        .order_by('-total_amount')

    for address_total in sn_amount_totals:
        address = address_total['clothes__sn']
        total_amount = address_total['total_amount']
        print(f"机件号: {address} - 数量: {total_amount}")
    he_amount_totals = Outbound.objects.values('clothes__sn') \
        .annotate(total_amount=Sum(F('clothes__price') * F('amount'))) \
        .order_by('-total_amount')

    for address_total in he_amount_totals:
        address = address_total['clothes__sn']
        total_amount = address_total['total_amount']
        print(f"小计: {address} - 数量: {total_amount}")









    context = {
        'address_amount_totals': address_amount_totals,
        'sn_amount_totals': sn_amount_totals,
        'he_amount_totals': he_amount_totals,



    }
    return render(request,'outbound/indexzd.html', context)
    return search(request, clothes=form.cleaned_data['clothes'], context=context)
def zdsearch(request, clothes=None, context=None):

    if request.method == "GET":
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')

        # 构建查询条件
        conditions = Q()

        if start_time and end_time:
            conditions &= Q(create_time__range=(start_time, end_time))

        address_amount_totals = Outbound.objects.filter(conditions).values('customer__address') \
            .annotate(total_amount=Sum('amount')) \
            .order_by('-total_amount')

        sn_amount_totals = Outbound.objects.filter(conditions).values('clothes__sn') \
            .annotate(total_amount=Sum('amount')) \
            .order_by('-total_amount')

        he_amount_totals = Outbound.objects.filter(conditions).values('clothes__sn') \
            .annotate(total_amount=Sum(F('clothes__price') * F('amount'))) \
            .order_by('-total_amount')



        context = {
            'address_amount_totals': address_amount_totals,
            'sn_amount_totals': sn_amount_totals,
            'he_amount_totals': he_amount_totals,

        }


        messages.add_message(request, messages.SUCCESS, '查询成功')
        return render(request, 'outbound/indexzd.html', context)
